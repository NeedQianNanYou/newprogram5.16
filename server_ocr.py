"""
听写批改 - 讯飞OCR版
HTTP传输避免SSL问题，r.encoding='utf-8'确保中文不乱码
"""
import base64, hashlib, io, json, os, re, time, uuid
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from PIL import Image
from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

env_path = Path(__file__).parent / ".env"
if env_path.exists():
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                if k.strip() not in os.environ:
                    os.environ[k.strip()] = v.strip()

XF_APPID = os.environ.get("XF_APPID", "")
XF_API_KEY = os.environ.get("XF_API_KEY", "")


def ocr(image_bytes):
    """讯飞手写OCR - 带重试"""
    b64 = base64.b64encode(image_bytes).decode()
    for attempt in range(3):
        cur = str(int(time.time()))
        param = '{"language":"cn|en","location":"false"}'
        pb64 = base64.b64encode(param.encode()).decode()
        cs = hashlib.md5((XF_API_KEY + cur + pb64).encode()).hexdigest()

        try:
            r = requests.post(
                "http://webapi.xfyun.cn/v1/service/v1/ocr/handwriting",
                headers={"X-CurTime": cur, "X-Param": pb64, "X-Appid": XF_APPID,
                         "X-CheckSum": cs},
                data={"image": b64}, timeout=30)
            r.encoding = "utf-8"
            result = r.json()
        except Exception as e:
            if attempt < 2:
                time.sleep(1)
                continue
            return None, str(e)

        if result.get("code") != "0":
            if attempt < 2:
                time.sleep(1)
                continue
            return None, result.get("desc", str(result))

        lines = []
        try:
            for line in result["data"]["block"][0]["line"]:
                text = "".join(w["content"] for w in line["word"])
                lines.append(text.strip())
        except (KeyError, IndexError):
            pass
        return lines, None

    return None, "重试3次均失败"


def parse_lines(lines):
    """解析OCR行 -> [{num, en, zh}]
    每行=1题。相邻'纯英文行+纯中文行'合并为1题"""
    raw = []
    for text in lines:
        text = text.strip()
        if not text:
            continue
        num_m = re.match(r'^(\d+)', text)
        num = int(num_m.group(1)) if num_m else 0
        en = " ".join(re.findall(r'[a-zA-Z\-]+', text))
        zh = "".join(re.findall(r'[一-鿿]+', text))
        raw.append({"num": num, "en": en, "zh": zh})

    # 合并相邻: 仅英文行 + 下一行仅中文行 -> 一题
    merged = []
    i = 0
    while i < len(raw):
        cur = raw[i]
        if cur["en"] and not cur["zh"] and i + 1 < len(raw):
            nxt = raw[i + 1]
            if nxt["zh"] and not nxt["en"]:
                merged.append({"num": cur["num"] or nxt["num"], "en": cur["en"], "zh": nxt["zh"]})
                i += 2
                continue
        merged.append(cur)
        i += 1

    pairs = []
    for m in merged:
        if m["en"] or m["zh"]:
            pairs.append({"num": m["num"], "en": m["en"], "zh": m["zh"]})
    return pairs


def img_to_jpg(img):
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=95)
    return buf.getvalue()


def split_columns(img, n=3):
    w, h = img.size
    cw = w // n
    return [img.crop((i * cw, 0, (i + 1) * cw if i < n - 1 else w, h)) for i in range(n)]


def ocr_column(col):
    lines, err = ocr(img_to_jpg(col))
    if err:
        raise RuntimeError(err)
    return parse_lines(lines)


def grade_paper(path, answer_key=None):
    img = Image.open(path)
    cols = split_columns(img, 3)

    all_pairs = []
    with ThreadPoolExecutor(max_workers=3) as pool:
        fs = {pool.submit(ocr_column, c): i for i, c in enumerate(cols)}
        for f in as_completed(fs):
            i = fs[f]
            try:
                pairs = f.result()
                offset = i * 33
                # 按列内位置重新编号：第1个=1, 第2个=2...加上列偏移
                for pos, p in enumerate(pairs, 1):
                    p["num"] = offset + pos
                print(f"  列{i+1}: OCR {len(pairs)}条 (题号{offset+1}-{offset+len(pairs)})")
                all_pairs.extend(pairs)
            except Exception as e:
                print(f"  列{i+1}失败: {e}")

    if answer_key:
        results = []
        for p in all_pairs:
            correct = answer_key.get(p["num"], "")
            student = p["zh"]
            if student and correct:
                judge = "对" if student.strip() == correct.strip() else "错"
            else:
                judge = "存疑"
            results.append({
                "题号": p["num"], "英文": p["en"],
                "正确翻译": correct or "(未知)",
                "学生写了": student or "(空白)", "判定": judge,
            })
        return results
    else:
        # 无词表：OCR读字 → AI文字判题（不传图片，不会编造）
        return judge_by_ai(all_pairs)


def judge_by_ai(pairs):
    """用千问文字模型判题——不看图片，只根据(英文,学生中文)文字对判断"""
    key = os.environ.get("DASHSCOPE_API_KEY", "")
    if not key:
        # 没有千问key，全标存疑
        return [{"题号": p["num"], "英文": p["en"],
                 "正确翻译": "(未指定)", "学生写了": p["zh"] or "(空白)",
                 "判定": "存疑"} for p in pairs]

    from openai import OpenAI
    client = OpenAI(api_key=key, base_url="https://dashscope.aliyuncs.com/compatible-mode/v1")

    items = "\n".join([f"{p['num']}. 英文:{p['en']}  学生写了:{p['zh']}" for p in pairs if p.get('en') and p.get('zh')])

    prompt = f"""把以下每个英文单词翻译成中文，然后判断学生写的中文对不对。

{items}

要求：
1. 先把每个英文单词翻译成中文（中文意思），写入 json 的 "中文意思" 字段
2. 判断学生写的对不对，写入 "判定" 字段

输出格式：
[{{"题号":1,"中文意思":"采取行动","判定":"对"}}, {{"题号":2,"中文意思":"前额","判定":"错"}}]

注意：
- "中文意思"字段只能填中文汉字，绝对不能填英文。例如填"苹果"而不是"apple"
- OCR可能把英文识别错了，结合学生写的中文来猜原单词
- 判对错时近义词算对

只返回JSON数组。"""

    resp = client.chat.completions.create(
        model="qwen-max", max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = resp.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = "\n".join(raw.split("\n")[1:-1])

    judgments = {}
    for item in json.loads(raw):
        judgments[item["题号"]] = item

    results = []
    for p in pairs:
        j = judgments.get(p["num"], {})
        corr = j.get("中文意思", j.get("正确翻译", ""))
        # AI可能返回英文而非中文——英文丢弃
        if corr and not re.search(r'[一-鿿]', corr):
            corr = ""
        results.append({
            "题号": p["num"], "英文": p["en"],
            "正确翻译": corr or "(未知)",
            "学生写了": p["zh"] or "(空白)",
            "判定": j.get("判定", "存疑"),
        })
    return results


def extract_key(path):
    img = Image.open(path)
    cols = split_columns(img, 3)
    key = {}

    def do(i, col):
        pairs = ocr_column(col)
        return {p.get("num", i*33): p["zh"] for p in pairs if p.get("zh")}

    with ThreadPoolExecutor(max_workers=3) as pool:
        fs = {pool.submit(do, i, c): i for i, c in enumerate(cols)}
        for f in as_completed(fs):
            try:
                key.update(f.result())
            except Exception:
                pass
    return key


def make_excel(all_results):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active; ws.title = "汇总"
    for c, h in enumerate(["学生", "总", "对", "错", "存疑", "正确率"], 1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    row = 2
    for name, results in sorted(all_results.items()):
        t = len(results)
        co = sum(1 for r in results if r["判定"] == "对")
        wr = sum(1 for r in results if r["判定"] == "错")
        un = sum(1 for r in results if r["判定"] == "存疑")
        for c, v in enumerate([name, t, co, wr, un, f"{co/t*100:.1f}%" if t else "0%"], 1):
            ws.cell(row=row, column=c, value=v)
        row += 1

    ws2 = wb.create_sheet("逐题")
    red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for c, h in enumerate(["学生", "题号", "英文", "正确翻译", "学生写了", "判定"], 1):
        ws2.cell(row=1, column=c, value=h).font = Font(bold=True)
    row = 2
    for name, results in sorted(all_results.items()):
        for r in sorted(results, key=lambda x: x["题号"]):
            vals = [name, r["题号"], r.get("英文", ""), r.get("正确翻译", ""), r.get("学生写了", ""), r["判定"]]
            for c, v in enumerate(vals, 1):
                ws2.cell(row=row, column=c, value=v)
            if r["判定"] == "错":
                for cc in range(1, 7):
                    ws2.cell(row=row, column=cc).fill = red
            row += 1

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out


app = FastAPI()
sd = Path(__file__).parent / "static"; sd.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(sd)), name="static")


@app.get("/")
async def index():
    return FileResponse(str(sd / "index_ocr.html"))


@app.post("/api/grade")
async def grade(files: list[UploadFile] = File(...)):
    try:
        ud = Path(__file__).parent / "uploads"; ud.mkdir(exist_ok=True)
        saved = []
        for f in files:
            ext = Path(f.filename).suffix or ".jpg"
            nm = f"{uuid.uuid4().hex[:8]}{ext}"
            p = ud / nm
            p.write_bytes(await f.read())
            saved.append((f.filename or "unknown", str(p)))

        key = None; st = 0
        if len(saved) >= 2:
            key = extract_key(saved[0][1]); st = 1
            print(f"词表: {len(key)}题")

        all_r = {}; errs = []

        def do(item):
            nm, p = item
            try:
                return Path(nm).stem, grade_paper(p, key), None
            except Exception as e:
                return Path(nm).stem, [], str(e)

        with ThreadPoolExecutor(max_workers=2) as pool:
            fs = [pool.submit(do, item) for item in saved[st:]]
            for f in as_completed(fs):
                nm, res, e = f.result()
                if e:
                    errs.append({"name": nm, "error": e})
                else:
                    all_r[nm] = res

        try:
            excel = make_excel(all_r)
            ep = ud / "批改结果.xlsx"
            ep.write_bytes(excel.read())
        except Exception as e:
            print(f"Excel: {e}")

        summary = []
        for name, results in sorted(all_r.items()):
            t = len(results)
            c = sum(1 for r in results if r["判定"] == "对")
            w = sum(1 for r in results if r["判定"] == "错")
            u = sum(1 for r in results if r["判定"] == "存疑")
            summary.append({"name": name, "total": t, "correct": c, "wrong": w, "uncertain": u,
                            "rate": f"{c/t*100:.1f}%" if t else "0%", "details": results})

        return JSONResponse({"success": True, "summary": summary, "errors": errs,
                             "answer_count": len(key) if key else 0, "excel_url": "/api/excel"})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JSONResponse({"success": False, "summary": [], "errors": [{"name": "系统", "error": str(e)}]}, status_code=200)


@app.get("/api/excel")
async def download():
    p = Path(__file__).parent / "uploads" / "批改结果.xlsx"
    return FileResponse(str(p), filename="批改结果.xlsx") if p.exists() else JSONResponse({"error": "无文件"}, 404)


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8765)
